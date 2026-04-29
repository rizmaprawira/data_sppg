#!/usr/bin/env python3
from __future__ import annotations

import argparse
import gc
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook, load_workbook
from shapely.geometry import box
from shapely.geometry.base import BaseGeometry
from shapely.ops import unary_union
from shapely.prepared import prep


SOURCE_WORKBOOK = Path("bgn_sppg_operasional.xlsx")
DEFAULT_OUTPUT_WORKBOOK = Path("bgn_sppg_operasional_geocoded_java_roads.xlsx")
DEFAULT_ADMIN_BOUNDARY_PATH = Path(
    "/Users/rizzie/Work/IndonesiaRe/data/batas_keldesa/Batas_Wilayah_KelurahanDesa_10K_AR.shp"
)
DEFAULT_ROAD_SHP_PATH = Path("/Users/rizzie/ClimateData/OSM_roads/gis_osm_roads_free_1.shp")

INPUT_COLUMNS = [
    "No",
    "Provinsi SPPG",
    "Kab./Kota SPPG",
    "Kecamatan SPPG",
    "Kelurahan/Desa SPPG",
    "Alamat SPPG",
    "Nama SPPG",
]

PROVINCE_GROUPS = {
    "sumatra": {
        "ACEH",
        "SUMATERA UTARA",
        "SUMATERA BARAT",
        "RIAU",
        "JAMBI",
        "SUMATERA SELATAN",
        "BENGKULU",
        "LAMPUNG",
        "KEPULAUAN RIAU",
        "KEPULAUAN BANGKA BELITUNG",
    },
    "java": {
        "BANTEN",
        "DKI JAKARTA",
        "JAKARTA",
        "JAWA BARAT",
        "JAWA TENGAH",
        "JAWA TIMUR",
        "DAERAH ISTIMEWA YOGYAKARTA",
        "YOGYAKARTA",
        "BALI",
    },
    "kalimantan": {
        "KALIMANTAN BARAT",
        "KALIMANTAN TENGAH",
        "KALIMANTAN SELATAN",
        "KALIMANTAN TIMUR",
        "KALIMANTAN UTARA",
    },
    "sulawesi": {
        "SULAWESI UTARA",
        "SULAWESI TENGAH",
        "SULAWESI SELATAN",
        "SULAWESI TENGGARA",
        "SULAWESI BARAT",
        "GORONTALO",
    },
    "nusa-tenggara": {
        "NUSA TENGGARA BARAT",
        "NUSA TENGGARA TIMUR",
    },
    "maluku": {
        "MALUKU",
        "MALUKU UTARA",
    },
    "papua": {
        "PAPUA",
        "PAPUA BARAT",
        "PAPUA BARAT DAYA",
        "PAPUA TENGAH",
        "PAPUA SELATAN",
        "PAPUA PEGUNUNGAN",
    },
}

DROP_PREFIX_TOKENS = {
    "JL",
    "JLN",
    "JALAN",
    "JALN",
    "JLAN",
    "GG",
    "GANG",
    "KOMP",
    "KOMPLEK",
    "KOMPLEKS",
    "PERUM",
    "PERUMAHAN",
    "BTN",
    "BLK",
    "BLOK",
    "RUKO",
    "RUKAN",
}

CANDIDATE_TERMINATORS = {
    "NO",
    "NOMOR",
    "RT",
    "RW",
    "KEL",
    "KELURAHAN",
    "KEC",
    "KECAMATAN",
    "KAB",
    "KABUPATEN",
    "PROV",
    "PROVINSI",
    "DEPAN",
    "BELAKANG",
    "SAMPING",
    "SEBELAH",
    "DALAM",
    "LUAR",
    "LINGKUNGAN",
    "LINGK",
    "KM",
}

ROAD_BLACKLIST = {
    "",
    "UNNAMED ROAD",
    "NO NAME",
    "TANPA NAMA",
    "JALAN TANPA NAMA",
    "UNKNOWN",
}

MAX_CANDIDATE_TOKENS = 8
MIN_MATCH_SCORE = 0.82
MIN_MATCH_MARGIN = 0.05


@dataclass(frozen=True)
class BoundaryCandidate:
    level: int
    source_index: int
    name_raw: str
    name_norm: str
    geometry: BaseGeometry
    score: float


@dataclass(frozen=True)
class AdminPathResult:
    province: BoundaryCandidate
    kabkota: BoundaryCandidate
    kecamatan: BoundaryCandidate | None = None
    kelurahan: BoundaryCandidate | None = None

    @property
    def boundary(self) -> BaseGeometry:
        if self.kelurahan is not None:
            return self.kelurahan.geometry
        if self.kecamatan is not None:
            return self.kecamatan.geometry
        return self.kabkota.geometry

    @property
    def admin_level_used(self) -> str:
        if self.kelurahan is not None:
            return "kelurahan"
        if self.kecamatan is not None:
            return "kecamatan"
        return "kabkota"


@dataclass(frozen=True)
class RoadMatchResult:
    success: bool
    latitude: float | None
    longitude: float | None
    road_name: str | None
    candidate_phrase: str | None
    score: float | None
    candidate_count: int
    reason: str


@dataclass(frozen=True)
class PreparedRoadLayer:
    road_geoms: dict[str, BaseGeometry]
    token_index: dict[str, set[str]]


@dataclass(frozen=True)
class JavaAdminCatalog:
    province_bounds: dict[str, tuple[float, float, float, float]]
    kabkota_bounds: dict[tuple[str, str], tuple[float, float, float, float]]
    kecamatan_bounds: dict[tuple[str, str, str], tuple[float, float, float, float]]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = text.replace("&", " AND ")
    text = re.sub(r"[/,;:()\-]", " ", text)
    text = re.sub(r"[.']", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def collapse_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def collapse_spaced_letters(text: str) -> str:
    tokens = text.split()
    if len(tokens) > 1 and all(len(token) == 1 for token in tokens):
        return "".join(tokens)
    return text


def normalize_admin_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = collapse_spaced_letters(text)
    text = re.sub(r"\bDAERAH ISTIMEWA\b", " ", text)
    text = re.sub(r"\bDAERAH KHUSUS IBUKOTA\b", " ", text)
    text = re.sub(r"\bPROVINSI\b", " ", text)
    text = re.sub(r"\bKABUPATEN\b", " ", text)
    text = re.sub(r"\bKAB\b", " ", text)
    text = re.sub(r"\bKOTA ADMINISTRASI\b", " ", text)
    text = re.sub(r"\bKOTA ADM\b", " ", text)
    text = re.sub(r"\bKOTA\b", " ", text)
    text = re.sub(r"\bKECAMATAN\b", " ", text)
    text = re.sub(r"\bKEC\b", " ", text)
    text = re.sub(r"\bKELURAHAN\b", " ", text)
    text = re.sub(r"\bKEL\b", " ", text)
    text = re.sub(r"\bDESA\b", " ", text)
    text = re.sub(r"\bDS\b", " ", text)
    text = re.sub(r"\bDSN\b", " ", text)
    text = collapse_spaces(text)
    return text


def normalize_kabkota_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = collapse_spaced_letters(text)
    text = re.sub(r"\bDAERAH ISTIMEWA\b", " ", text)
    text = re.sub(r"\bDAERAH KHUSUS IBUKOTA\b", " ", text)
    text = re.sub(r"\bPROVINSI\b", " ", text)
    text = re.sub(r"\bKABUPATEN\b", " ", text)
    text = re.sub(r"\bKAB\b", " ", text)
    text = re.sub(r"\bKOTA ADMINISTRASI\b", " ", text)
    text = re.sub(r"\bKOTA ADM\b", " ", text)
    text = re.sub(r"\bKOTA\b", " ", text)
    text = re.sub(r"\bADM\b", " ", text)
    text = re.sub(r"\bADMINISTRASI\b", " ", text)
    text = re.sub(r"\bKECAMATAN\b", " ", text)
    text = re.sub(r"\bKEC\b", " ", text)
    text = re.sub(r"\bKELURAHAN\b", " ", text)
    text = re.sub(r"\bKEL\b", " ", text)
    text = re.sub(r"\bDESA\b", " ", text)
    text = re.sub(r"\bDS\b", " ", text)
    text = re.sub(r"\bDSN\b", " ", text)
    text = collapse_spaces(text)
    if text in {"KEP SERIBU", "KEPULAUAN SERIBU", "SERIBU"}:
        return "KEPULAUAN SERIBU"
    return text


def detect_admin_kind(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return "unknown"
    if re.search(r"\bKOTA\b", text) or "KOTA ADM" in text:
        return "city"
    if re.search(r"\bKABUPATEN\b", text) or re.search(r"\bKAB\b", text):
        return "regency"
    return "unknown"


def parse_admin_level(value: object) -> int | None:
    if value is None:
        return None
    text = normalize_text(value)
    match = re.search(r"\d+", text)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def road_prefix_tokens(tokens: list[str]) -> list[str]:
    if not tokens:
        return tokens
    start = 0
    while start < len(tokens) and tokens[start] in DROP_PREFIX_TOKENS:
        start += 1
    if start < len(tokens) and tokens[start] == "RAYA" and start > 0:
        start += 1
    return tokens[start:]


def normalize_road_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    tokens = road_prefix_tokens(text.split())
    text = collapse_spaces(" ".join(tokens))
    if text in ROAD_BLACKLIST:
        return ""
    return text


def tokenize_for_index(value: str) -> tuple[str, ...]:
    tokens = []
    for token in value.split():
        if token in DROP_PREFIX_TOKENS:
            continue
        if token in CANDIDATE_TERMINATORS:
            continue
        if len(token) < 3 and not token.isdigit():
            continue
        tokens.append(token)
    return tuple(tokens)


def aliases_for_text(value: object) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    aliases = {raw, normalize_admin_name(raw)}
    return [alias for alias in aliases if alias]


def candidate_phrases(address: object) -> list[str]:
    text = normalize_text(address)
    if not text:
        return []

    segments = [segment for segment in re.split(r"[,\n;|]+", text) if segment.strip()]
    if not segments:
        segments = [text]

    phrases: list[str] = []
    for segment in segments:
        tokens = segment.split()
        if not tokens:
            continue

        variants = [tokens]
        if tokens and tokens[0] in DROP_PREFIX_TOKENS:
            variants.append(tokens[1:])
            if len(tokens) > 1 and tokens[1] == "RAYA":
                variants.append(tokens[2:])
        if tokens and tokens[0] == "RAYA":
            variants.append(tokens[1:])

        for variant in variants:
            variant = list(variant)
            if not variant:
                continue
            cutoff = len(variant)
            for idx, token in enumerate(variant):
                if token in CANDIDATE_TERMINATORS:
                    cutoff = idx
                    break
            variant = variant[:cutoff]
            if not variant:
                continue
            max_tokens = min(len(variant), MAX_CANDIDATE_TOKENS)
            for end in range(max_tokens, 0, -1):
                phrase = collapse_spaces(" ".join(variant[:end]))
                if phrase:
                    phrases.append(phrase)

    ordered = list(dict.fromkeys(phrases))
    ordered = [phrase for _, phrase in sorted(enumerate(ordered), key=lambda item: (-len(item[1].split()), item[0]))]
    return ordered


def compact_text(text: str) -> str:
    return text.replace(" ", "")


def road_similarity(candidate: str, road_name: str) -> float:
    candidate = normalize_road_name(candidate)
    road_name = normalize_road_name(road_name)
    if not candidate or not road_name:
        return 0.0

    if candidate == road_name:
        return 1.0

    candidate_compact = compact_text(candidate)
    road_compact = compact_text(road_name)
    if candidate_compact == road_compact:
        return 1.0
    if candidate_compact in road_compact or road_compact in candidate_compact:
        return 0.98

    candidate_tokens = set(tokenize_for_index(candidate))
    road_tokens = set(tokenize_for_index(road_name))
    if candidate_tokens and road_tokens:
        overlap = len(candidate_tokens & road_tokens)
        token_score = overlap / max(len(candidate_tokens), len(road_tokens))
    else:
        token_score = 0.0

    sequence_score = SequenceMatcher(None, candidate_compact, road_compact).ratio()
    return max(sequence_score, token_score)


@lru_cache(maxsize=1)
def province_group_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for group, provinces in PROVINCE_GROUPS.items():
        for province in provinces:
            lookup[normalize_admin_name(province)] = group
    return lookup


def determine_extract_group(province: object) -> str:
    province_norm = normalize_admin_name(province)
    return province_group_lookup().get(province_norm, "sumatra")


def build_java_admin_catalog(admin_boundary_path: Path) -> JavaAdminCatalog:
    boundaries = gpd.read_file(
        admin_boundary_path,
        columns=["WADMPR", "WADMKK", "WADMKC"],
    )
    boundaries = boundaries[boundaries.geometry.notna()].copy()
    boundaries["province_key"] = boundaries["WADMPR"].map(normalize_admin_name)
    boundaries["kabkota_key"] = boundaries["WADMKK"].map(normalize_kabkota_name)
    boundaries["kecamatan_key"] = boundaries["WADMKC"].map(normalize_admin_name)
    boundaries = boundaries[
        boundaries["province_key"].ne("")
        & boundaries["kabkota_key"].ne("")
        & boundaries["kecamatan_key"].ne("")
    ].copy()

    province_bounds: dict[str, tuple[float, float, float, float]] = {}
    for key, group in boundaries.groupby("province_key", sort=False):
        province_bounds[key] = tuple(float(value) for value in group.geometry.total_bounds)

    kabkota_bounds: dict[tuple[str, str], tuple[float, float, float, float]] = {}
    for key, group in boundaries.groupby(["province_key", "kabkota_key"], sort=False):
        kabkota_bounds[key] = tuple(float(value) for value in group.geometry.total_bounds)

    kecamatan_bounds: dict[tuple[str, str, str], tuple[float, float, float, float]] = {}
    for key, group in boundaries.groupby(["province_key", "kabkota_key", "kecamatan_key"], sort=False):
        kecamatan_bounds[key] = tuple(float(value) for value in group.geometry.total_bounds)

    return JavaAdminCatalog(
        province_bounds=province_bounds,
        kabkota_bounds=kabkota_bounds,
        kecamatan_bounds=kecamatan_bounds,
    )


def _make_boundary_candidate(
    level: int,
    name_raw: str,
    name_norm: str,
    geometry: BaseGeometry,
) -> BoundaryCandidate:
    return BoundaryCandidate(
        level=level,
        source_index=-1,
        name_raw=name_raw,
        name_norm=name_norm,
        geometry=geometry,
        score=1.0,
    )


def java_row_keys(row: pd.Series | dict[str, object]) -> tuple[str, str, str, str]:
    province_key = normalize_admin_name(row["Provinsi SPPG"])
    kabkota_key = normalize_kabkota_name(row["Kab./Kota SPPG"])
    kecamatan_key = normalize_admin_name(row["Kecamatan SPPG"])
    kelurahan_key = normalize_admin_name(row["Kelurahan/Desa SPPG"])
    return province_key, kabkota_key, kecamatan_key, kelurahan_key


def java_job_key(row: pd.Series | dict[str, object]) -> tuple[str, str, str]:
    province_key, kabkota_key, kecamatan_key, _ = java_row_keys(row)
    return province_key, kabkota_key, kecamatan_key


def java_job_geometry(
    catalog: JavaAdminCatalog,
    job_key: tuple[str, str, str],
) -> BaseGeometry | None:
    province_key, kabkota_key, kecamatan_key = job_key

    bounds = catalog.kecamatan_bounds.get((province_key, kabkota_key, kecamatan_key))
    if bounds is not None:
        return box(*bounds)

    bounds = catalog.kabkota_bounds.get((province_key, kabkota_key))
    if bounds is not None:
        return box(*bounds)

    bounds = catalog.province_bounds.get(province_key)
    if bounds is not None:
        return box(*bounds)
    return None


def build_java_admin_path(
    row: pd.Series | dict[str, object],
    catalog: JavaAdminCatalog,
    group_geometry: BaseGeometry,
) -> AdminPathResult:
    province_key, kabkota_key, kecamatan_key, kelurahan_key = java_row_keys(row)
    province_raw = str(row["Provinsi SPPG"])
    kabkota_raw = str(row["Kab./Kota SPPG"])
    kecamatan_raw = str(row["Kecamatan SPPG"])
    kelurahan_raw = str(row["Kelurahan/Desa SPPG"])

    kec_bounds = catalog.kecamatan_bounds.get((province_key, kabkota_key, kecamatan_key))
    if kec_bounds is not None:
        geometry = box(*kec_bounds)
        province = _make_boundary_candidate(4, province_raw, province_key, geometry)
        kabkota = _make_boundary_candidate(5, kabkota_raw, kabkota_key, geometry)
        kecamatan = _make_boundary_candidate(6, kecamatan_raw, kecamatan_key, geometry)
        kelurahan = _make_boundary_candidate(7, kelurahan_raw, kelurahan_key, geometry)
        return AdminPathResult(province=province, kabkota=kabkota, kecamatan=kecamatan, kelurahan=kelurahan)

    province = _make_boundary_candidate(4, province_raw, province_key, group_geometry)
    kabkota = _make_boundary_candidate(5, kabkota_raw, kabkota_key, group_geometry)
    kecamatan = _make_boundary_candidate(6, kecamatan_raw, kecamatan_key, group_geometry)
    kelurahan = _make_boundary_candidate(7, kelurahan_raw, kelurahan_key, group_geometry)
    return AdminPathResult(province=province, kabkota=kabkota, kecamatan=kecamatan, kelurahan=kelurahan)


def load_input_rows(workbook_path: Path) -> tuple[str, pd.DataFrame]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Workbook {workbook_path} is empty")
        header = list(rows[0])
        data = pd.DataFrame(rows[1:], columns=header)
        data["source_excel_row"] = range(2, len(data) + 2)
        return ws.title, data
    finally:
        wb.close()


def output_sheet_columns() -> list[str]:
    address_index = INPUT_COLUMNS.index("Alamat SPPG")
    return INPUT_COLUMNS[:address_index] + ["latitude", "longitude"] + INPUT_COLUMNS[address_index:]


def unresolved_sheet_headers() -> list[str]:
    return [
        "source_excel_row",
        "No",
        "Provinsi SPPG",
        "Kab./Kota SPPG",
        "Kecamatan SPPG",
        "Kelurahan/Desa SPPG",
        "Alamat SPPG",
        "Nama SPPG",
        "status",
        "reason",
        "admin_level_used",
        "candidate_phrase",
        "matched_road_name",
        "match_score",
        "candidate_count",
    ]


def output_row_values(row: dict[str, object], result: dict[str, object] | None = None) -> list[object]:
    latitude = None if result is None else result["latitude"]
    longitude = None if result is None else result["longitude"]
    return [
        row["No"],
        row["Provinsi SPPG"],
        row["Kab./Kota SPPG"],
        row["Kecamatan SPPG"],
        row["Kelurahan/Desa SPPG"],
        latitude,
        longitude,
        row["Alamat SPPG"],
        row["Nama SPPG"],
    ]


def unresolved_row_values(row: dict[str, object], result: dict[str, object]) -> list[object]:
    return [
        row["source_excel_row"],
        row["No"],
        row["Provinsi SPPG"],
        row["Kab./Kota SPPG"],
        row["Kecamatan SPPG"],
        row["Kelurahan/Desa SPPG"],
        row["Alamat SPPG"],
        row["Nama SPPG"],
        result["status"],
        result["reason"],
        result["admin_level_used"],
        result["candidate_phrase"],
        result["matched_road_name"],
        result["match_score"],
        result["candidate_count"],
    ]


def create_output_workbook(rows: pd.DataFrame, sheet_name: str) -> tuple[Workbook, object, object, dict[int, int]]:
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name
    review_ws = workbook.create_sheet("unresolved_rows")

    ws.append(output_sheet_columns())
    review_ws.append(unresolved_sheet_headers())
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    row_map: dict[int, int] = {}
    for output_row_num, row in enumerate(rows.to_dict(orient="records"), start=2):
        source_excel_row = int(row["source_excel_row"])
        row_map[source_excel_row] = output_row_num
        ws.append(output_row_values(row))

    return workbook, ws, review_ws, row_map


def apply_results_to_workbook(
    ws,
    review_ws,
    rows: pd.DataFrame,
    results: list[dict[str, object]],
    row_map: dict[int, int],
) -> None:
    row_records = rows.to_dict(orient="records")
    if len(row_records) != len(results):
        raise ValueError(f"Row/result mismatch: {len(row_records)} rows vs {len(results)} results")

    for row, result in zip(row_records, results):
        output_row_num = row_map[int(row["source_excel_row"])]
        ws.cell(row=output_row_num, column=6, value=result["latitude"])
        ws.cell(row=output_row_num, column=7, value=result["longitude"])
        if result["status"] != "matched":
            review_ws.append(unresolved_row_values(row, result))


def save_output_workbook(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_output_workbook(
    rows: pd.DataFrame,
    results: list[dict[str, object]],
    output_path: Path,
    sheet_name: str,
) -> None:
    workbook, ws, review_ws, row_map = create_output_workbook(rows, sheet_name)
    apply_results_to_workbook(ws, review_ws, rows, results, row_map)
    save_output_workbook(workbook, output_path)


def build_road_index(roads: gpd.GeoDataFrame) -> PreparedRoadLayer:
    road_geoms: dict[str, BaseGeometry] = {}
    token_index: dict[str, set[str]] = defaultdict(set)

    grouped = roads.groupby("road_name_norm", sort=False)
    for road_name, group in grouped:
        geometries = [geom for geom in group.geometry if geom is not None and not geom.is_empty]
        if not geometries:
            continue
        if len(geometries) == 1:
            merged = geometries[0]
        else:
            merged = unary_union(geometries)
        if merged.is_empty:
            continue
        road_geoms[road_name] = merged
        tokens = tokenize_for_index(road_name)
        for token in tokens:
            token_index[token].add(road_name)

    return PreparedRoadLayer(road_geoms=road_geoms, token_index=token_index)


def load_road_layer_for_bbox(road_shp_path: Path, bbox: Iterable[float]) -> PreparedRoadLayer:
    bbox = tuple(bbox)
    roads = gpd.read_file(
        road_shp_path,
        bbox=bbox,
        columns=["name"],
    )
    if not isinstance(roads, gpd.GeoDataFrame):
        roads = gpd.GeoDataFrame(roads)
    roads = roads[roads.geometry.notna()].copy()
    if "name" not in roads.columns:
        return PreparedRoadLayer(road_geoms={}, token_index={})
    roads = roads[roads["name"].notna()].copy()
    roads["name_raw"] = roads["name"].astype(str)
    roads["road_name_norm"] = roads["name_raw"].map(normalize_road_name)
    roads = roads[roads["road_name_norm"].ne("")].copy()
    roads = roads[~roads["road_name_norm"].isin(ROAD_BLACKLIST)].copy()
    roads = roads.reset_index(drop=True)
    return build_road_index(roads)


def road_name_candidates(
    roads: PreparedRoadLayer,
    phrase: str,
) -> list[tuple[str, float, int]]:
    phrase_norm = normalize_road_name(phrase)
    if not phrase_norm:
        return []

    exact = []
    if phrase_norm in roads.road_geoms:
        exact.append((phrase_norm, 1.0, 10_000))
    if exact:
        return exact

    tokens = tokenize_for_index(phrase_norm)
    if not tokens:
        return []

    candidate_counts: Counter[str] = Counter()
    for token in tokens:
        for road_name in roads.token_index.get(token, set()):
            candidate_counts[road_name] += 1

    if not candidate_counts:
        return []

    min_overlap = 1 if len(tokens) == 1 else 2
    candidates = [road_name for road_name, count in candidate_counts.items() if count >= min_overlap]
    if not candidates:
        candidates = [name for name, _ in candidate_counts.most_common(100)]

    scored: list[tuple[str, float, int]] = []
    for road_name in candidates:
        score = road_similarity(phrase_norm, road_name)
        if score <= 0:
            continue
        scored.append((road_name, score, candidate_counts[road_name]))

    scored.sort(key=lambda item: (-item[1], -item[2], item[0]))
    return scored


def choose_road_geometry(
    roads: PreparedRoadLayer,
    road_name: str,
    boundary_geometry: BaseGeometry,
) -> BaseGeometry | None:
    road_geometry = roads.road_geoms.get(road_name)
    if road_geometry is None or road_geometry.is_empty:
        return None
    try:
        clipped = road_geometry.intersection(boundary_geometry)
    except Exception:
        return None
    if clipped.is_empty:
        return None
    return clipped


def geocode_with_path(
    roads: PreparedRoadLayer,
    path: AdminPathResult,
    address: object,
) -> RoadMatchResult:
    phrases = candidate_phrases(address)
    if not phrases:
        return RoadMatchResult(False, None, None, None, None, None, 0, "no_address_candidate")

    boundary_geometry = path.boundary
    best: tuple[str, str, float, int, BaseGeometry] | None = None
    candidate_total = 0

    for phrase in phrases:
        scored_candidates = road_name_candidates(roads, phrase)
        candidate_total += len(scored_candidates)
        if not scored_candidates:
            continue

        top_score = scored_candidates[0][1]
        second_score = scored_candidates[1][1] if len(scored_candidates) > 1 else 0.0
        if top_score < MIN_MATCH_SCORE or (top_score - second_score) < MIN_MATCH_MARGIN:
            continue

        road_name, score, overlap = scored_candidates[0]
        clipped = choose_road_geometry(roads, road_name, boundary_geometry)
        if clipped is None:
            continue
        best = (phrase, road_name, score, overlap, clipped)
        break

    if best is None:
        return RoadMatchResult(False, None, None, None, None, None, candidate_total, "road_candidate_not_matched")

    phrase, road_name, score, overlap, clipped = best
    point = clipped.representative_point()
    return RoadMatchResult(
        True,
        float(point.y),
        float(point.x),
        road_name,
        phrase,
        score,
        candidate_total,
        "matched",
    )


def geocode_row(
    roads: PreparedRoadLayer,
    row: pd.Series | dict[str, object],
    admin_path: AdminPathResult,
) -> dict[str, object]:
    road_result = geocode_with_path(roads, admin_path, row["Alamat SPPG"])
    if road_result.success:
        return {
            "latitude": road_result.latitude,
            "longitude": road_result.longitude,
            "status": "matched",
            "reason": "matched",
            "admin_level_used": admin_path.admin_level_used,
            "best_candidate": road_result.road_name,
            "matched_road_name": road_result.road_name,
            "candidate_phrase": road_result.candidate_phrase,
            "match_score": road_result.score,
            "candidate_count": road_result.candidate_count,
        }

    return {
        "latitude": None,
        "longitude": None,
        "status": "unresolved",
        "reason": road_result.reason,
        "admin_level_used": admin_path.admin_level_used,
        "best_candidate": None,
        "matched_road_name": None,
        "candidate_phrase": None,
        "match_score": None,
        "candidate_count": road_result.candidate_count,
    }


def process_java_rows(df: pd.DataFrame, args: argparse.Namespace, sheet_name: str) -> None:
    catalog = build_java_admin_catalog(args.admin_boundaries)

    work_df = df.copy()
    work_df["java_job_key"] = [java_job_key(row) for row in work_df.to_dict(orient="records")]

    workbook, ws, review_ws, row_map = create_output_workbook(work_df, sheet_name)
    save_output_workbook(workbook, args.output)
    print(f"[group] initialized output workbook at {args.output.name}")

    matched_total = 0
    unresolved_total = 0
    processed_total = 0

    job_order = list(dict.fromkeys(work_df["java_job_key"].tolist()))
    for job_key in job_order:
        job_rows = work_df.loc[work_df["java_job_key"] == job_key]
        job_geometry = java_job_geometry(catalog, job_key)
        if job_geometry is None:
            raise ValueError(f"Could not resolve a Java geometry for group {job_key}")

        bbox = tuple(job_geometry.bounds)
        province_key, kabkota_key, kecamatan_key = job_key
        print(f"[group] {province_key} / {kabkota_key} / {kecamatan_key} rows={len(job_rows)} bbox={bbox}")
        print(f"[roads] loading {args.roads.name} for {province_key} / {kabkota_key} / {kecamatan_key}")

        roads = load_road_layer_for_bbox(args.roads, bbox=bbox)
        group_results = []
        for row in job_rows.to_dict(orient="records"):
            admin_path = build_java_admin_path(row, catalog, job_geometry)
            group_results.append(geocode_row(roads, row, admin_path=admin_path))

        apply_results_to_workbook(ws, review_ws, job_rows, group_results, row_map)

        group_matched = sum(1 for item in group_results if item["status"] == "matched")
        group_unresolved = len(group_results) - group_matched
        matched_total += group_matched
        unresolved_total += group_unresolved
        processed_total += len(group_results)

        save_output_workbook(workbook, args.output)
        print(
            f"[group] saved {args.output.name} after {province_key} / {kabkota_key} / {kecamatan_key}: "
            f"processed={processed_total}/{len(work_df)} matched={matched_total} unresolved={unresolved_total}"
        )

        del roads
        del group_results
        gc.collect()

    print(f"[done] matched={matched_total} unresolved={unresolved_total}")
    print(f"[done] wrote {args.output.resolve()}")


def default_output_path() -> Path:
    return DEFAULT_OUTPUT_WORKBOOK


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Geocode Java SPPG workbook using a local roads shapefile.")
    parser.add_argument(
        "--input",
        type=Path,
        default=SOURCE_WORKBOOK,
        help="Input XLSX workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX workbook path.",
    )
    parser.add_argument(
        "--roads",
        type=Path,
        default=DEFAULT_ROAD_SHP_PATH,
        help="Road shapefile path derived from OSM.",
    )
    parser.add_argument(
        "--admin-boundaries",
        type=Path,
        default=DEFAULT_ADMIN_BOUNDARY_PATH,
        help="Java kelurahan boundary shapefile path.",
    )
    args = parser.parse_args(argv)

    sheet_name, df = load_input_rows(args.input)
    missing = [column for column in INPUT_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"Workbook {args.input} is missing expected columns: {missing}")

    df = df.copy()
    df["extract_group"] = df["Provinsi SPPG"].map(determine_extract_group)
    df = df.loc[df["extract_group"] == "java"].copy()
    if df.empty:
        raise ValueError("No Java rows matched the input workbook.")
    df = df.loc[:, INPUT_COLUMNS + ["source_excel_row"]].copy()

    if args.output is None:
        args.output = default_output_path()

    process_java_rows(df, args, sheet_name)


if __name__ == "__main__":
    main()
