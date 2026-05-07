#!/usr/bin/env python3
from __future__ import annotations

import argparse
import gc
import hashlib
import random
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable, Iterable, Sequence

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook, load_workbook
from shapely.geometry import GeometryCollection, LineString, MultiLineString, Point
from shapely.geometry.base import BaseGeometry
from shapely.ops import unary_union


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
SOURCE_WORKBOOK = PROJECT_DIR / "bgn_sppg_operasional.xlsx"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"
DEFAULT_ADMIN_BOUNDARY_PATH = Path(
    "/Users/rizzie/Work/IndonesiaRe/data/batas_keldesa/Batas_Wilayah_KelurahanDesa_10K_AR.shp"
)
DEFAULT_ROAD_SHP_PATH = Path("/Users/rizzie/ClimateData/OSM_roads/gis_osm_roads_free_1.shp")

INPUT_COLUMNS = [
    "No",
    "Provinsi SPPG",
    "Kab./Kota SPPG",
    "Kecamatan SPPG",
    "Kelurahan/Desa SPPG",
    "Alamat SPPG",
    "Nama SPPG",
]

PROVINCE_GROUPS = {
    "sumatra": {
        "ACEH",
        "SUMATERA UTARA",
        "SUMATERA BARAT",
        "RIAU",
        "JAMBI",
        "SUMATERA SELATAN",
        "BENGKULU",
        "LAMPUNG",
        "KEPULAUAN RIAU",
        "KEPULAUAN BANGKA BELITUNG",
    },
    "java": {
        "BANTEN",
        "DKI JAKARTA",
        "JAKARTA",
        "JAWA BARAT",
        "JAWA TENGAH",
        "JAWA TIMUR",
        "DAERAH ISTIMEWA YOGYAKARTA",
        "YOGYAKARTA",
        "BALI",
    },
    "kalimantan": {
        "KALIMANTAN BARAT",
        "KALIMANTAN TENGAH",
        "KALIMANTAN SELATAN",
        "KALIMANTAN TIMUR",
        "KALIMANTAN UTARA",
    },
    "sulawesi": {
        "SULAWESI UTARA",
        "SULAWESI TENGAH",
        "SULAWESI SELATAN",
        "SULAWESI TENGGARA",
        "SULAWESI BARAT",
        "GORONTALO",
    },
    "nusa-tenggara": {
        "NUSA TENGGARA BARAT",
        "NUSA TENGGARA TIMUR",
    },
    "maluku": {
        "MALUKU",
        "MALUKU UTARA",
    },
    "papua": {
        "PAPUA",
        "PAPUA BARAT",
        "PAPUA BARAT DAYA",
        "PAPUA TENGAH",
        "PAPUA SELATAN",
        "PAPUA PEGUNUNGAN",
    },
}

DROP_PREFIX_TOKENS = {
    "JL",
    "JLN",
    "JALAN",
    "JALN",
    "JLAN",
    "GG",
    "GANG",
    "KOMP",
    "KOMPLEK",
    "KOMPLEKS",
    "PERUM",
    "PERUMAHAN",
    "BTN",
    "BLK",
    "BLOK",
    "RUKO",
    "RUKAN",
}

CANDIDATE_TERMINATORS = {
    "NO",
    "NOMOR",
    "RT",
    "RW",
    "KEL",
    "KELURAHAN",
    "KEC",
    "KECAMATAN",
    "KAB",
    "KABUPATEN",
    "PROV",
    "PROVINSI",
    "DEPAN",
    "BELAKANG",
    "SAMPING",
    "SEBELAH",
    "DALAM",
    "LUAR",
    "LINGKUNGAN",
    "LINGK",
    "KM",
}

ROAD_BLACKLIST = {
    "",
    "UNNAMED ROAD",
    "NO NAME",
    "TANPA NAMA",
    "JALAN TANPA NAMA",
    "UNKNOWN",
}

MAX_CANDIDATE_TOKENS = 8
MIN_MATCH_SCORE = 0.82
MIN_MATCH_MARGIN = 0.05

MATCH_BY_JALAN = "jalan"
MATCH_BY_KELURAHAN = "kelurahan/desa"
MATCH_BY_KECAMATAN = "kecamatan"
MATCH_BY_KABKOTA = "kabupaten/kota"


@dataclass(frozen=True)
class AreaGeometry:
    key: tuple[str, ...]
    name: str
    geometry: BaseGeometry

    @property
    def bounds(self) -> tuple[float, float, float, float]:
        return tuple(float(value) for value in self.geometry.bounds)


@dataclass(frozen=True)
class RoadMatchResult:
    success: bool
    latitude: float | None
    longitude: float | None
    road_name: str | None
    candidate_phrase: str | None
    score: float | None
    candidate_count: int
    reason: str


@dataclass
class PreparedRoadLayer:
    roads: gpd.GeoDataFrame
    road_geoms: dict[str, BaseGeometry]
    token_index: dict[str, set[str]]
    sampled_geometry_cache: dict[tuple[str, ...], list[tuple[BaseGeometry, float]]]


@dataclass(frozen=True)
class AdminCatalog:
    province_geoms: dict[tuple[str], AreaGeometry]
    kabkota_geoms: dict[tuple[str, str], AreaGeometry]
    kecamatan_geoms: dict[tuple[str, str, str], AreaGeometry]
    kelurahan_geoms: dict[tuple[str, str, str, str], AreaGeometry]
    provinces: tuple[str, ...]
    kabkota_by_province: dict[tuple[str], tuple[tuple[str, str], ...]]
    kecamatan_by_kabkota: dict[tuple[str, str], tuple[tuple[str, str, str], ...]]
    kelurahan_by_kecamatan: dict[tuple[str, str, str], tuple[tuple[str, str, str, str], ...]]
    kelurahan_by_kabkota: dict[tuple[str, str], tuple[tuple[str, str, str, str], ...]]


@dataclass(frozen=True)
class GeocodeResult:
    latitude: float | None
    longitude: float | None
    match_by: str | None
    status: str
    reason: str
    matched_road_name: str | None
    candidate_phrase: str | None
    match_score: float | None
    candidate_count: int


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = text.replace("&", " AND ")
    text = re.sub(r"[/,;:()\-]", " ", text)
    text = re.sub(r"[.']", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def collapse_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def collapse_spaced_letters(text: str) -> str:
    tokens = text.split()
    if len(tokens) > 1 and all(len(token) == 1 for token in tokens):
        return "".join(tokens)
    return text


def compact_text(text: str) -> str:
    return text.replace(" ", "")


def normalize_province_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = collapse_spaced_letters(text)
    text = re.sub(r"\bPROV\b", " ", text)
    text = re.sub(r"\bPROVINSI\b", " ", text)
    text = re.sub(r"\bDAERAH ISTIMEWA\b", " ", text)
    text = re.sub(r"\bDAERAH KHUSUS IBUKOTA\b", " ", text)
    return collapse_spaces(text)


def normalize_admin_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = collapse_spaced_letters(text)
    text = re.sub(r"\bDAERAH ISTIMEWA\b", " ", text)
    text = re.sub(r"\bDAERAH KHUSUS IBUKOTA\b", " ", text)
    text = re.sub(r"\bPROVINSI\b", " ", text)
    text = re.sub(r"\bKABUPATEN\b", " ", text)
    text = re.sub(r"\bKAB\b", " ", text)
    text = re.sub(r"\bKOTA ADMINISTRASI\b", " ", text)
    text = re.sub(r"\bKOTA ADM\b", " ", text)
    text = re.sub(r"\bKOTA\b", " ", text)
    text = re.sub(r"\bKECAMATAN\b", " ", text)
    text = re.sub(r"\bKEC\b", " ", text)
    text = re.sub(r"\bKELURAHAN\b", " ", text)
    text = re.sub(r"\bKEL\b", " ", text)
    text = re.sub(r"\bDESA\b", " ", text)
    text = re.sub(r"\bDS\b", " ", text)
    text = re.sub(r"\bDSN\b", " ", text)
    return collapse_spaces(text)


def normalize_kabkota_name(value: object) -> str:
    text = normalize_admin_name(value)
    text = re.sub(r"\bADM\b", " ", text)
    text = re.sub(r"\bADMINISTRASI\b", " ", text)
    text = collapse_spaces(text)
    if text in {"KEP SERIBU", "KEPULAUAN SERIBU", "SERIBU"}:
        return "KEPULAUAN SERIBU"
    return text


def determine_extract_group(province: object) -> str:
    province_norm = normalize_province_name(province)
    for group_name, provinces in PROVINCE_GROUPS.items():
        if province_norm in provinces:
            return group_name
    return "other"


def road_prefix_tokens(tokens: list[str]) -> list[str]:
    if not tokens:
        return tokens
    start = 0
    while start < len(tokens) and tokens[start] in DROP_PREFIX_TOKENS:
        start += 1
    if start < len(tokens) and tokens[start] == "RAYA" and start > 0:
        start += 1
    return tokens[start:]


def normalize_road_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    tokens = road_prefix_tokens(text.split())
    text = collapse_spaces(" ".join(tokens))
    if text in ROAD_BLACKLIST:
        return ""
    return text


def tokenize_for_index(value: str) -> tuple[str, ...]:
    tokens = []
    for token in value.split():
        if token in DROP_PREFIX_TOKENS:
            continue
        if token in CANDIDATE_TERMINATORS:
            continue
        if len(token) < 3 and not token.isdigit():
            continue
        tokens.append(token)
    return tuple(tokens)


def candidate_phrases(address: object) -> list[str]:
    text = normalize_text(address)
    if not text:
        return []

    segments = [segment for segment in re.split(r"[,\n;|]+", text) if segment.strip()]
    if not segments:
        segments = [text]

    phrases: list[str] = []
    for segment in segments:
        tokens = segment.split()
        if not tokens:
            continue

        variants = [tokens]
        if tokens and tokens[0] in DROP_PREFIX_TOKENS:
            variants.append(tokens[1:])
            if len(tokens) > 1 and tokens[1] == "RAYA":
                variants.append(tokens[2:])
        if tokens and tokens[0] == "RAYA":
            variants.append(tokens[1:])

        for variant in variants:
            variant = list(variant)
            if not variant:
                continue
            cutoff = len(variant)
            for idx, token in enumerate(variant):
                if token in CANDIDATE_TERMINATORS:
                    cutoff = idx
                    break
            variant = variant[:cutoff]
            if not variant:
                continue
            max_tokens = min(len(variant), MAX_CANDIDATE_TOKENS)
            for end in range(max_tokens, 0, -1):
                phrase = collapse_spaces(" ".join(variant[:end]))
                if phrase:
                    phrases.append(phrase)

    ordered = list(dict.fromkeys(phrases))
    ordered = [phrase for _, phrase in sorted(enumerate(ordered), key=lambda item: (-len(item[1].split()), item[0]))]
    return ordered


def road_similarity(candidate: str, road_name: str) -> float:
    candidate = normalize_road_name(candidate)
    road_name = normalize_road_name(road_name)
    if not candidate or not road_name:
        return 0.0
    if candidate == road_name:
        return 1.0

    candidate_compact = compact_text(candidate)
    road_compact = compact_text(road_name)
    if candidate_compact == road_compact:
        return 1.0
    if candidate_compact in road_compact or road_compact in candidate_compact:
        return 0.98

    candidate_tokens = set(tokenize_for_index(candidate))
    road_tokens = set(tokenize_for_index(road_name))
    token_score = 0.0
    if candidate_tokens and road_tokens:
        token_score = len(candidate_tokens & road_tokens) / max(len(candidate_tokens), len(road_tokens))

    sequence_score = SequenceMatcher(None, candidate_compact, road_compact).ratio()
    return max(sequence_score, token_score)


def _union_geometries(geometries: Iterable[BaseGeometry]) -> BaseGeometry | None:
    values = [geometry for geometry in geometries if geometry is not None and not geometry.is_empty]
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    merged = unary_union(values)
    if merged.is_empty:
        return None
    return merged


def _collect_linear_parts(geometry: BaseGeometry | None) -> list[BaseGeometry]:
    if geometry is None or geometry.is_empty:
        return []
    if isinstance(geometry, LineString):
        return [geometry] if geometry.length > 0 else []
    if isinstance(geometry, MultiLineString):
        return [part for part in geometry.geoms if part is not None and not part.is_empty and part.length > 0]
    if isinstance(geometry, GeometryCollection):
        parts: list[BaseGeometry] = []
        for part in geometry.geoms:
            parts.extend(_collect_linear_parts(part))
        return parts
    return []


def _point_on_linear_geometry(geometry: BaseGeometry, fraction: float) -> Point:
    parts = _collect_linear_parts(geometry)
    if not parts:
        return geometry.representative_point()

    total_length = sum(part.length for part in parts)
    if total_length <= 0:
        return parts[0].representative_point()

    target = total_length * min(max(fraction, 0.0), 1.0)
    walked = 0.0
    for part in parts:
        next_walked = walked + part.length
        if target <= next_walked:
            distance = max(target - walked, 0.0)
            return part.interpolate(distance)
        walked = next_walked
    return parts[-1].interpolate(parts[-1].length)


def _random_point_on_linear_geometry(geometry: BaseGeometry, rng: random.Random) -> Point:
    parts = _collect_linear_parts(geometry)
    if not parts:
        return geometry.representative_point()

    total_length = sum(part.length for part in parts)
    if total_length <= 0:
        return parts[0].representative_point()

    target = rng.random() * total_length
    walked = 0.0
    for part in parts:
        next_walked = walked + part.length
        if target <= next_walked:
            return part.interpolate(max(target - walked, 0.0))
        walked = next_walked
    return parts[-1].interpolate(parts[-1].length)


def stable_row_rng(row: dict[str, object], level: str, base_seed: int) -> random.Random:
    payload = "|".join(
        [
            str(base_seed),
            level,
            str(row.get("source_excel_row", "")),
            str(row.get("No", "")),
            str(row.get("Provinsi SPPG", "")),
            str(row.get("Kab./Kota SPPG", "")),
            str(row.get("Kecamatan SPPG", "")),
            str(row.get("Kelurahan/Desa SPPG", "")),
            str(row.get("Nama SPPG", "")),
        ]
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    return random.Random(int(digest[:16], 16))


def load_input_rows(workbook_path: Path) -> tuple[str, pd.DataFrame]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Workbook {workbook_path} is empty")
        header = list(rows[0])
        data = pd.DataFrame(rows[1:], columns=header)
        data["source_excel_row"] = range(2, len(data) + 2)
        return ws.title, data
    finally:
        wb.close()


def build_admin_catalog(boundary_path: Path, island: str) -> AdminCatalog:
    boundaries = gpd.read_file(
        boundary_path,
        columns=["WADMPR", "WADMKK", "WADMKC", "WADMKD", "geometry"],
    )
    boundaries = boundaries[boundaries.geometry.notna()].copy()
    boundaries["province_key"] = boundaries["WADMPR"].map(normalize_province_name)
    boundaries["kabkota_key"] = boundaries["WADMKK"].map(normalize_kabkota_name)
    boundaries["kecamatan_key"] = boundaries["WADMKC"].map(normalize_admin_name)
    boundaries["kelurahan_key"] = boundaries["WADMKD"].map(normalize_admin_name)
    boundaries = boundaries[
        boundaries["province_key"].isin(PROVINCE_GROUPS[island])
        & boundaries["province_key"].ne("")
        & boundaries["kabkota_key"].ne("")
        & boundaries["kecamatan_key"].ne("")
        & boundaries["kelurahan_key"].ne("")
    ].copy()
    boundaries = boundaries.reset_index(drop=True)

    province_geoms: dict[tuple[str], AreaGeometry] = {}
    for province_key, group in boundaries.groupby("province_key", sort=False):
        geometry = _union_geometries(group.geometry)
        if geometry is None:
            continue
        key = (province_key,)
        province_geoms[key] = AreaGeometry(key=key, name=str(group.iloc[0]["WADMPR"]), geometry=geometry)

    kabkota_geoms: dict[tuple[str, str], AreaGeometry] = {}
    for key, group in boundaries.groupby(["province_key", "kabkota_key"], sort=False):
        geometry = _union_geometries(group.geometry)
        if geometry is None:
            continue
        kabkota_geoms[key] = AreaGeometry(key=key, name=str(group.iloc[0]["WADMKK"]), geometry=geometry)

    kecamatan_geoms: dict[tuple[str, str, str], AreaGeometry] = {}
    for key, group in boundaries.groupby(["province_key", "kabkota_key", "kecamatan_key"], sort=False):
        geometry = _union_geometries(group.geometry)
        if geometry is None:
            continue
        kecamatan_geoms[key] = AreaGeometry(key=key, name=str(group.iloc[0]["WADMKC"]), geometry=geometry)

    kelurahan_geoms: dict[tuple[str, str, str, str], AreaGeometry] = {}
    for key, group in boundaries.groupby(["province_key", "kabkota_key", "kecamatan_key", "kelurahan_key"], sort=False):
        geometry = _union_geometries(group.geometry)
        if geometry is None:
            continue
        kelurahan_geoms[key] = AreaGeometry(key=key, name=str(group.iloc[0]["WADMKD"]), geometry=geometry)

    kabkota_by_province: dict[tuple[str], list[tuple[str, str]]] = defaultdict(list)
    for key in kabkota_geoms:
        kabkota_by_province[(key[0],)].append(key)

    kecamatan_by_kabkota: dict[tuple[str, str], list[tuple[str, str, str]]] = defaultdict(list)
    for key in kecamatan_geoms:
        kecamatan_by_kabkota[(key[0], key[1])].append(key)

    kelurahan_by_kecamatan: dict[tuple[str, str, str], list[tuple[str, str, str, str]]] = defaultdict(list)
    kelurahan_by_kabkota: dict[tuple[str, str], list[tuple[str, str, str, str]]] = defaultdict(list)
    for key in kelurahan_geoms:
        kelurahan_by_kecamatan[(key[0], key[1], key[2])].append(key)
        kelurahan_by_kabkota[(key[0], key[1])].append(key)

    return AdminCatalog(
        province_geoms=province_geoms,
        kabkota_geoms=kabkota_geoms,
        kecamatan_geoms=kecamatan_geoms,
        kelurahan_geoms=kelurahan_geoms,
        provinces=tuple(key[0] for key in province_geoms.keys()),
        kabkota_by_province={key: tuple(value) for key, value in kabkota_by_province.items()},
        kecamatan_by_kabkota={key: tuple(value) for key, value in kecamatan_by_kabkota.items()},
        kelurahan_by_kecamatan={key: tuple(value) for key, value in kelurahan_by_kecamatan.items()},
        kelurahan_by_kabkota={key: tuple(value) for key, value in kelurahan_by_kabkota.items()},
    )


def resolve_single_key(
    query: object,
    candidates: Sequence[str],
    normalizer: Callable[[object], str],
    min_score: float = 0.75,
) -> str | None:
    query_norm = normalizer(query)
    if not query_norm or not candidates:
        return None
    if query_norm in candidates:
        return query_norm

    compact_query = compact_text(query_norm)
    compact_matches = [candidate for candidate in candidates if compact_text(candidate) == compact_query]
    if len(compact_matches) == 1:
        return compact_matches[0]

    containment_matches = [
        candidate
        for candidate in candidates
        if compact_query and (compact_query in compact_text(candidate) or compact_text(candidate) in compact_query)
    ]
    if len(containment_matches) == 1:
        return containment_matches[0]

    best_key: str | None = None
    best_score = 0.0
    for candidate in candidates:
        score = SequenceMatcher(None, compact_query, compact_text(candidate)).ratio()
        if compact_query and compact_text(candidate):
            if compact_query in compact_text(candidate) or compact_text(candidate) in compact_query:
                score = max(score, 0.99)
        if score > best_score:
            best_score = score
            best_key = candidate
    if best_key is not None and best_score >= min_score:
        return best_key
    return None


def resolve_composite_key(
    query: object,
    candidates: Sequence[tuple[str, ...]],
    normalizer: Callable[[object], str],
    value_getter: Callable[[tuple[str, ...]], str],
    min_score: float = 0.75,
) -> tuple[str, ...] | None:
    if not candidates:
        return None
    names = [value_getter(candidate) for candidate in candidates]
    resolved_name = resolve_single_key(query, names, normalizer=normalizer, min_score=min_score)
    if resolved_name is None:
        return None
    matched = [candidate for candidate in candidates if value_getter(candidate) == resolved_name]
    if not matched:
        return None
    if len(matched) == 1:
        return matched[0]
    return sorted(matched)[0]


def resolve_province_key(catalog: AdminCatalog, province: object) -> tuple[str] | None:
    province_name = resolve_single_key(province, catalog.provinces, normalize_province_name)
    if province_name is None:
        return None
    return (province_name,)


def resolve_kabkota_key(catalog: AdminCatalog, province_key: tuple[str] | None, kabkota: object) -> tuple[str, str] | None:
    if province_key is None:
        return None
    candidates = catalog.kabkota_by_province.get(province_key, ())
    return resolve_composite_key(
        kabkota,
        candidates,
        normalizer=normalize_kabkota_name,
        value_getter=lambda key: key[1],
    )


def resolve_kecamatan_key(
    catalog: AdminCatalog,
    kabkota_key: tuple[str, str] | None,
    kecamatan: object,
) -> tuple[str, str, str] | None:
    if kabkota_key is None:
        return None
    candidates = catalog.kecamatan_by_kabkota.get(kabkota_key, ())
    return resolve_composite_key(
        kecamatan,
        candidates,
        normalizer=normalize_admin_name,
        value_getter=lambda key: key[2],
    )


def resolve_kelurahan_key(
    catalog: AdminCatalog,
    kabkota_key: tuple[str, str] | None,
    kecamatan_key: tuple[str, str, str] | None,
    kelurahan: object,
) -> tuple[str, str, str, str] | None:
    if kecamatan_key is not None:
        candidates = catalog.kelurahan_by_kecamatan.get(kecamatan_key, ())
    elif kabkota_key is not None:
        candidates = catalog.kelurahan_by_kabkota.get(kabkota_key, ())
    else:
        candidates = ()
    return resolve_composite_key(
        kelurahan,
        candidates,
        normalizer=normalize_admin_name,
        value_getter=lambda key: key[3],
    )


def build_road_index(roads: gpd.GeoDataFrame) -> PreparedRoadLayer:
    road_geoms: dict[str, BaseGeometry] = {}
    token_index: dict[str, set[str]] = defaultdict(set)
    grouped = roads.groupby("road_name_norm", sort=False)
    for road_name, group in grouped:
        geometry = _union_geometries(group.geometry)
        if geometry is None:
            continue
        road_geoms[road_name] = geometry
        for token in tokenize_for_index(road_name):
            token_index[token].add(road_name)
    return PreparedRoadLayer(
        roads=roads,
        road_geoms=road_geoms,
        token_index=token_index,
        sampled_geometry_cache={},
    )


def load_road_layer_for_bbox(road_shp_path: Path, bbox: Iterable[float]) -> PreparedRoadLayer:
    roads = gpd.read_file(road_shp_path, bbox=tuple(bbox), columns=["name"])
    if not isinstance(roads, gpd.GeoDataFrame):
        roads = gpd.GeoDataFrame(roads)
    roads = roads[roads.geometry.notna()].copy()
    if "name" not in roads.columns:
        roads["name"] = None
    roads = roads.reset_index(drop=True)

    named_roads = roads[roads["name"].notna()].copy()
    named_roads["road_name_norm"] = named_roads["name"].astype(str).map(normalize_road_name)
    named_roads = named_roads[named_roads["road_name_norm"].ne("")].copy()
    named_roads = named_roads[~named_roads["road_name_norm"].isin(ROAD_BLACKLIST)].copy()
    named_roads = named_roads.reset_index(drop=True)

    prepared = build_road_index(named_roads)
    prepared.roads = roads
    return prepared


def road_name_candidates(roads: PreparedRoadLayer, phrase: str) -> list[tuple[str, float, int]]:
    phrase_norm = normalize_road_name(phrase)
    if not phrase_norm:
        return []
    if phrase_norm in roads.road_geoms:
        return [(phrase_norm, 1.0, 10_000)]

    tokens = tokenize_for_index(phrase_norm)
    if not tokens:
        return []

    candidate_counts: Counter[str] = Counter()
    for token in tokens:
        for road_name in roads.token_index.get(token, set()):
            candidate_counts[road_name] += 1
    if not candidate_counts:
        return []

    min_overlap = 1 if len(tokens) == 1 else 2
    candidates = [road_name for road_name, count in candidate_counts.items() if count >= min_overlap]
    if not candidates:
        candidates = [road_name for road_name, _count in candidate_counts.most_common(100)]

    scored: list[tuple[str, float, int]] = []
    for road_name in candidates:
        score = road_similarity(phrase_norm, road_name)
        if score <= 0:
            continue
        scored.append((road_name, score, candidate_counts[road_name]))
    scored.sort(key=lambda item: (-item[1], -item[2], item[0]))
    return scored


def choose_road_geometry(roads: PreparedRoadLayer, road_name: str, boundary_geometry: BaseGeometry) -> BaseGeometry | None:
    road_geometry = roads.road_geoms.get(road_name)
    if road_geometry is None or road_geometry.is_empty:
        return None
    try:
        clipped = road_geometry.intersection(boundary_geometry)
    except Exception:
        return None
    if clipped.is_empty:
        return None
    if not _collect_linear_parts(clipped):
        return None
    return clipped


def geocode_by_jalan(
    roads: PreparedRoadLayer,
    boundary_geometry: BaseGeometry,
    address: object,
) -> RoadMatchResult:
    phrases = candidate_phrases(address)
    if not phrases:
        return RoadMatchResult(False, None, None, None, None, None, 0, "no_address_candidate")

    best: tuple[str, str, float, int, BaseGeometry] | None = None
    candidate_total = 0
    for phrase in phrases:
        scored_candidates = road_name_candidates(roads, phrase)
        candidate_total += len(scored_candidates)
        if not scored_candidates:
            continue

        top_score = scored_candidates[0][1]
        second_score = scored_candidates[1][1] if len(scored_candidates) > 1 else 0.0
        if top_score < MIN_MATCH_SCORE or (top_score - second_score) < MIN_MATCH_MARGIN:
            continue

        road_name, score, overlap = scored_candidates[0]
        clipped = choose_road_geometry(roads, road_name, boundary_geometry)
        if clipped is None:
            continue
        best = (phrase, road_name, score, overlap, clipped)
        break

    if best is None:
        return RoadMatchResult(False, None, None, None, None, None, candidate_total, "road_candidate_not_matched")

    phrase, road_name, score, overlap, clipped = best
    point = _point_on_linear_geometry(clipped, 0.5)
    return RoadMatchResult(
        True,
        latitude=float(point.y),
        longitude=float(point.x),
        road_name=road_name,
        candidate_phrase=phrase,
        score=score,
        candidate_count=candidate_total + overlap,
        reason="matched",
    )


def clipped_road_geometries_for_area(
    roads: PreparedRoadLayer,
    area: AreaGeometry,
) -> list[tuple[BaseGeometry, float]]:
    cache_key = area.key
    cached = roads.sampled_geometry_cache.get(cache_key)
    if cached is not None:
        return cached

    if roads.roads.empty:
        roads.sampled_geometry_cache[cache_key] = []
        return []

    candidate_indices = list(roads.roads.sindex.intersection(area.bounds))
    sampled: list[tuple[BaseGeometry, float]] = []
    for idx in candidate_indices:
        geometry = roads.roads.geometry.iloc[idx]
        if geometry is None or geometry.is_empty:
            continue
        try:
            clipped = geometry.intersection(area.geometry)
        except Exception:
            continue
        for part in _collect_linear_parts(clipped):
            if part.length > 0:
                sampled.append((part, float(part.length)))

    roads.sampled_geometry_cache[cache_key] = sampled
    return sampled


def sample_random_road_point(
    roads: PreparedRoadLayer,
    area: AreaGeometry,
    rng: random.Random,
) -> Point | None:
    candidates = clipped_road_geometries_for_area(roads, area)
    if not candidates:
        return None

    total_length = sum(length for _geometry, length in candidates)
    if total_length <= 0:
        return candidates[0][0].representative_point()

    target = rng.random() * total_length
    walked = 0.0
    for geometry, length in candidates:
        next_walked = walked + length
        if target <= next_walked:
            local_fraction = 0.0 if length == 0 else (target - walked) / length
            return _point_on_linear_geometry(geometry, local_fraction)
        walked = next_walked
    return _random_point_on_linear_geometry(candidates[-1][0], rng)


def build_job_key(row: pd.Series | dict[str, object]) -> tuple[str, str]:
    return (
        normalize_province_name(row["Provinsi SPPG"]),
        normalize_kabkota_name(row["Kab./Kota SPPG"]),
    )


def geocode_row(
    roads: PreparedRoadLayer,
    catalog: AdminCatalog,
    row: dict[str, object],
    base_seed: int,
) -> GeocodeResult:
    province_key = resolve_province_key(catalog, row["Provinsi SPPG"])
    kabkota_key = resolve_kabkota_key(catalog, province_key, row["Kab./Kota SPPG"])
    if province_key is None or kabkota_key is None:
        return GeocodeResult(
            latitude=None,
            longitude=None,
            match_by=None,
            status="unresolved",
            reason="kabupaten_kota_boundary_not_found",
            matched_road_name=None,
            candidate_phrase=None,
            match_score=None,
            candidate_count=0,
        )

    kabkota_area = catalog.kabkota_geoms.get(kabkota_key)
    if kabkota_area is None:
        return GeocodeResult(
            latitude=None,
            longitude=None,
            match_by=None,
            status="unresolved",
            reason="kabupaten_kota_geometry_missing",
            matched_road_name=None,
            candidate_phrase=None,
            match_score=None,
            candidate_count=0,
        )

    jalan_result = geocode_by_jalan(roads, kabkota_area.geometry, row["Alamat SPPG"])
    if jalan_result.success:
        return GeocodeResult(
            latitude=jalan_result.latitude,
            longitude=jalan_result.longitude,
            match_by=MATCH_BY_JALAN,
            status="matched",
            reason="matched_by_jalan",
            matched_road_name=jalan_result.road_name,
            candidate_phrase=jalan_result.candidate_phrase,
            match_score=jalan_result.score,
            candidate_count=jalan_result.candidate_count,
        )

    kecamatan_key = resolve_kecamatan_key(catalog, kabkota_key, row["Kecamatan SPPG"])
    kelurahan_key = resolve_kelurahan_key(catalog, kabkota_key, kecamatan_key, row["Kelurahan/Desa SPPG"])

    area_chain: list[tuple[str, AreaGeometry | None, str]] = [
        (MATCH_BY_KELURAHAN, catalog.kelurahan_geoms.get(kelurahan_key) if kelurahan_key else None, "kelurahan_no_road"),
        (MATCH_BY_KECAMATAN, catalog.kecamatan_geoms.get(kecamatan_key) if kecamatan_key else None, "kecamatan_no_road"),
        (MATCH_BY_KABKOTA, kabkota_area, "kabupaten_kota_no_road"),
    ]

    for match_by, area, failure_reason in area_chain:
        if area is None:
            continue
        rng = stable_row_rng(row, match_by, base_seed)
        point = sample_random_road_point(roads, area, rng)
        if point is None:
            continue
        return GeocodeResult(
            latitude=float(point.y),
            longitude=float(point.x),
            match_by=match_by,
            status="matched",
            reason=f"matched_by_{match_by}",
            matched_road_name=None,
            candidate_phrase=None,
            match_score=None,
            candidate_count=jalan_result.candidate_count,
        )

    if kelurahan_key is None and kecamatan_key is None:
        reason = "kelurahan_dan_kecamatan_boundary_not_found"
    elif kelurahan_key is None:
        reason = "kelurahan_boundary_not_found"
    elif kecamatan_key is None:
        reason = "kecamatan_boundary_not_found"
    else:
        reason = "no_road_found_in_boundaries"

    return GeocodeResult(
        latitude=None,
        longitude=None,
        match_by=None,
        status="unresolved",
        reason=reason,
        matched_road_name=None,
        candidate_phrase=jalan_result.candidate_phrase,
        match_score=jalan_result.score,
        candidate_count=jalan_result.candidate_count,
    )


def output_sheet_columns() -> list[str]:
    return [
        "No",
        "Provinsi SPPG",
        "Kab./Kota SPPG",
        "Kecamatan SPPG",
        "Kelurahan/Desa SPPG",
        "Alamat SPPG",
        "Nama SPPG",
        "Latitude",
        "Longitude",
        "Match_By",
    ]


def unresolved_sheet_headers() -> list[str]:
    return [
        "source_excel_row",
        "No",
        "Provinsi SPPG",
        "Kab./Kota SPPG",
        "Kecamatan SPPG",
        "Kelurahan/Desa SPPG",
        "Alamat SPPG",
        "Nama SPPG",
        "Latitude",
        "Longitude",
        "Match_By",
        "status",
        "reason",
        "candidate_phrase",
        "matched_road_name",
        "match_score",
        "candidate_count",
    ]


def output_row_values(row: dict[str, object], result: GeocodeResult | None = None) -> list[object]:
    latitude = None if result is None else result.latitude
    longitude = None if result is None else result.longitude
    match_by = None if result is None else result.match_by
    return [
        row["No"],
        row["Provinsi SPPG"],
        row["Kab./Kota SPPG"],
        row["Kecamatan SPPG"],
        row["Kelurahan/Desa SPPG"],
        row["Alamat SPPG"],
        row["Nama SPPG"],
        latitude,
        longitude,
        match_by,
    ]


def unresolved_row_values(row: dict[str, object], result: GeocodeResult) -> list[object]:
    return [
        row["source_excel_row"],
        row["No"],
        row["Provinsi SPPG"],
        row["Kab./Kota SPPG"],
        row["Kecamatan SPPG"],
        row["Kelurahan/Desa SPPG"],
        row["Alamat SPPG"],
        row["Nama SPPG"],
        result.latitude,
        result.longitude,
        result.match_by,
        result.status,
        result.reason,
        result.candidate_phrase,
        result.matched_road_name,
        result.match_score,
        result.candidate_count,
    ]


def create_output_workbook(rows: pd.DataFrame, sheet_name: str) -> tuple[Workbook, object, object, dict[int, int]]:
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name
    review_ws = workbook.create_sheet("unresolved_rows")

    ws.append(output_sheet_columns())
    review_ws.append(unresolved_sheet_headers())
    row_map: dict[int, int] = {}
    for output_row_num, row in enumerate(rows.to_dict(orient="records"), start=2):
        source_excel_row = int(row["source_excel_row"])
        row_map[source_excel_row] = output_row_num
        ws.append(output_row_values(row))

    return workbook, ws, review_ws, row_map


def save_output_workbook(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def load_existing_output_workbook(output_path: Path) -> tuple[Workbook, object, object] | None:
    if not output_path.exists():
        return None
    workbook = load_workbook(output_path)
    if not workbook.sheetnames:
        return None
    ws = workbook[workbook.sheetnames[0]]
    review_ws = workbook["unresolved_rows"] if "unresolved_rows" in workbook.sheetnames else workbook.create_sheet("unresolved_rows")
    return workbook, ws, review_ws


def existing_processed_source_rows(rows: pd.DataFrame, ws, review_ws) -> set[int]:
    processed: set[int] = set()
    row_records = rows.to_dict(orient="records")
    for output_row_num, row in enumerate(row_records, start=2):
        lat = ws.cell(row=output_row_num, column=8).value
        lon = ws.cell(row=output_row_num, column=9).value
        match_by = ws.cell(row=output_row_num, column=10).value
        if lat is not None and lon is not None and match_by:
            processed.add(int(row["source_excel_row"]))

    if review_ws is not None:
        for review_row in review_ws.iter_rows(min_row=2, values_only=True):
            if not review_row or review_row[0] is None:
                continue
            try:
                processed.add(int(review_row[0]))
            except (TypeError, ValueError):
                continue
    return processed


def existing_progress_counts(rows: pd.DataFrame, ws, review_ws) -> tuple[int, int]:
    matched = 0
    for output_row_num, _row in enumerate(rows.to_dict(orient="records"), start=2):
        lat = ws.cell(row=output_row_num, column=8).value
        lon = ws.cell(row=output_row_num, column=9).value
        match_by = ws.cell(row=output_row_num, column=10).value
        if lat is not None and lon is not None and match_by:
            matched += 1
    unresolved = 0
    if review_ws is not None:
        unresolved = max(review_ws.max_row - 1, 0)
    return matched, unresolved


def prepare_output_state(
    rows: pd.DataFrame,
    sheet_name: str,
    output_path: Path,
    restart: bool = False,
) -> tuple[Workbook, object, object, dict[int, int], set[int]]:
    existing = None if restart else load_existing_output_workbook(output_path)
    if existing is None:
        workbook, ws, review_ws, row_map = create_output_workbook(rows, sheet_name)
        save_output_workbook(workbook, output_path)
        return workbook, ws, review_ws, row_map, set()

    workbook, ws, review_ws = existing
    row_map = {
        int(row["source_excel_row"]): output_row_num
        for output_row_num, row in enumerate(rows.to_dict(orient="records"), start=2)
    }
    processed_rows = existing_processed_source_rows(rows, ws, review_ws)
    return workbook, ws, review_ws, row_map, processed_rows


def apply_results_to_workbook(
    ws,
    review_ws,
    rows: pd.DataFrame,
    results: list[GeocodeResult],
    row_map: dict[int, int],
) -> None:
    row_records = rows.to_dict(orient="records")
    if len(row_records) != len(results):
        raise ValueError(f"Row/result mismatch: {len(row_records)} rows vs {len(results)} results")

    for row, result in zip(row_records, results):
        output_row_num = row_map[int(row["source_excel_row"])]
        ws.cell(row=output_row_num, column=8, value=result.latitude)
        ws.cell(row=output_row_num, column=9, value=result.longitude)
        ws.cell(row=output_row_num, column=10, value=result.match_by)
        if result.status != "matched":
            review_ws.append(unresolved_row_values(row, result))


def process_island_rows(df: pd.DataFrame, args: argparse.Namespace, sheet_name: str) -> None:
    catalog = build_admin_catalog(args.admin_boundaries, args.island)

    work_df = df.copy()
    work_df["job_key"] = [build_job_key(row) for row in work_df.to_dict(orient="records")]

    workbook, ws, review_ws, row_map, processed_rows = prepare_output_state(
        work_df,
        sheet_name,
        args.output,
        restart=args.restart,
    )

    if processed_rows:
        print(f"[group] resuming {args.output.name} with {len(processed_rows)} already processed rows")
    else:
        print(f"[group] initialized output workbook at {args.output.name}")

    matched_total, unresolved_total = existing_progress_counts(work_df, ws, review_ws)
    processed_total = len(processed_rows)

    job_order = list(dict.fromkeys(work_df["job_key"].tolist()))
    for job_key in job_order:
        job_rows = work_df.loc[(work_df["job_key"] == job_key) & (~work_df["source_excel_row"].isin(processed_rows))]
        if job_rows.empty:
            continue

        province_norm, kabkota_norm = job_key
        province_key = resolve_province_key(catalog, province_norm)
        kabkota_key = resolve_kabkota_key(catalog, province_key, kabkota_norm)
        if province_key is None or kabkota_key is None or kabkota_key not in catalog.kabkota_geoms:
            group_results = [
                GeocodeResult(
                    latitude=None,
                    longitude=None,
                    match_by=None,
                    status="unresolved",
                    reason="kabupaten_kota_boundary_not_found",
                    matched_road_name=None,
                    candidate_phrase=None,
                    match_score=None,
                    candidate_count=0,
                )
                for _ in range(len(job_rows))
            ]
            apply_results_to_workbook(ws, review_ws, job_rows, group_results, row_map)
            processed_rows.update(int(row["source_excel_row"]) for row in job_rows.to_dict(orient="records"))
            unresolved_total += len(group_results)
            processed_total += len(group_results)
            save_output_workbook(workbook, args.output)
            print(
                f"[group] saved {args.output.name} after unresolved {province_norm} / {kabkota_norm}: "
                f"processed={processed_total}/{len(work_df)} matched={matched_total} unresolved={unresolved_total}"
            )
            continue

        kabkota_area = catalog.kabkota_geoms[kabkota_key]
        bbox = kabkota_area.bounds
        print(f"[group] {province_norm} / {kabkota_norm} rows={len(job_rows)} bbox={bbox}")
        print(f"[roads] loading {args.roads.name} for {province_norm} / {kabkota_norm}")

        roads = load_road_layer_for_bbox(args.roads, bbox=bbox)
        group_results = [
            geocode_row(roads, catalog, row, base_seed=args.seed)
            for row in job_rows.to_dict(orient="records")
        ]

        apply_results_to_workbook(ws, review_ws, job_rows, group_results, row_map)
        processed_rows.update(int(row["source_excel_row"]) for row in job_rows.to_dict(orient="records"))

        group_matched = sum(1 for item in group_results if item.status == "matched")
        group_unresolved = len(group_results) - group_matched
        matched_total += group_matched
        unresolved_total += group_unresolved
        processed_total += len(group_results)

        save_output_workbook(workbook, args.output)
        print(
            f"[group] saved {args.output.name} after {province_norm} / {kabkota_norm}: "
            f"processed={processed_total}/{len(work_df)} matched={matched_total} unresolved={unresolved_total}"
        )

        del roads
        del group_results
        gc.collect()

    print(f"[done] island={args.island} matched={matched_total} unresolved={unresolved_total}")
    print(f"[done] wrote {args.output.resolve()}")


def default_output_path(island: str) -> Path:
    return DEFAULT_OUTPUT_DIR / f"bgn_sppg_operasional_geocoded_{island}_v2_roads_levels.xlsx"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Geocode SPPG workbook using road-name matching and road-based admin fallbacks.")
    parser.add_argument(
        "--island",
        choices=sorted(PROVINCE_GROUPS.keys()),
        required=True,
        help="Island group to process.",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=SOURCE_WORKBOOK,
        help="Input XLSX workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output XLSX workbook path.",
    )
    parser.add_argument(
        "--roads",
        type=Path,
        default=DEFAULT_ROAD_SHP_PATH,
        help="Road shapefile path derived from OSM.",
    )
    parser.add_argument(
        "--admin-boundaries",
        type=Path,
        default=DEFAULT_ADMIN_BOUNDARY_PATH,
        help="Kelurahan/desa boundary shapefile path.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=20260507,
        help="Base seed for deterministic random road sampling.",
    )
    parser.add_argument(
        "--restart",
        action="store_true",
        help="Ignore any existing output workbook and start a fresh run.",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)

    sheet_name, df = load_input_rows(args.input)
    missing = [column for column in INPUT_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"Workbook {args.input} is missing expected columns: {missing}")

    df = df.copy()
    df["extract_group"] = df["Provinsi SPPG"].map(determine_extract_group)
    df = df.loc[df["extract_group"] == args.island].copy()
    if df.empty:
        raise ValueError(f"No rows matched island group: {args.island}")
    df = df.loc[:, INPUT_COLUMNS + ["source_excel_row"]].copy()

    if args.output is None:
        args.output = default_output_path(args.island)

    process_island_rows(df, args, sheet_name)


if __name__ == "__main__":
    main()
